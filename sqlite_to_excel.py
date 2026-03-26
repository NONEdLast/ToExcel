#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
SQLite数据库与Excel表格相互转换工具
功能：
1. 将SQLite数据库中的一个或多个表转换为Excel文件
2. 将Excel文件中的数据转换为SQLite数据库表
3. 支持计算Excel中的函数并将结果存储到SQLite
4. 支持选择特定表进行转换
5. 支持Unicode排序
"""

import sqlite3
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formula.translate import Translator
from openpyxl import load_workbook
import os


def sqlite_to_excel(db_path, excel_path, tables=None, sort_by_unicode=False):
    """
    将SQLite数据库转换为Excel文件
    
    参数：
    db_path: SQLite数据库文件路径
    excel_path: 输出Excel文件路径
    tables: 要转换的表名列表，None表示转换所有表
    sort_by_unicode: 是否按Unicode排序字段名
    
    返回：
    bool: 转换是否成功
    """
    try:
        print(f"开始SQLite到Excel的转换...")
        print(f"输入SQLite文件：{db_path}")
        print(f"输出Excel文件：{excel_path}")
        
        # 检查SQLite文件是否存在
        if not os.path.exists(db_path):
            print(f"错误：SQLite文件 {db_path} 不存在！")
            return False
            
        # 检查SQLite文件是否可读
        if not os.access(db_path, os.R_OK):
            print(f"错误：无法读取SQLite文件 {db_path}！")
            return False
            
        # 检查输出目录是否存在
        output_dir = os.path.dirname(excel_path)
        if output_dir and not os.path.exists(output_dir):
            try:
                os.makedirs(output_dir)
                print(f"已创建输出目录：{output_dir}")
            except Exception as e:
                print(f"错误：无法创建输出目录 {output_dir} - {str(e)}")
                return False
                
        # 检查输出文件是否可写
        if os.path.exists(excel_path):
            if not os.access(excel_path, os.W_OK):
                print(f"错误：无法写入Excel文件 {excel_path}！")
                return False
        else:
            if output_dir and not os.access(output_dir, os.W_OK):
                print(f"错误：无法写入输出目录 {output_dir}！")
                return False
        
        # 连接SQLite数据库
        print("连接SQLite数据库...")
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        
        # 获取所有表名
        print("获取数据库中的表名...")
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%';")
        all_tables = [table[0] for table in cursor.fetchall()]
        print(f"数据库中的表：{all_tables}")
        
        # 确定要转换的表
        tables_to_convert = tables if tables is not None else all_tables
        print(f"要转换的表：{tables_to_convert}")
        tables_to_convert = [table for table in tables_to_convert if table in all_tables]
        print(f"最终转换的表：{tables_to_convert}")
        
        if not tables_to_convert:
            print("错误：没有找到要转换的表")
            conn.close()
            return False
        
        # 创建Excel工作簿
        print("创建Excel工作簿...")
        writer = pd.ExcelWriter(excel_path, engine='openpyxl')
        
        for table_name in tables_to_convert:
            # 读取表数据
            print(f"读取表 {table_name} 的数据...")
            query = f"SELECT * FROM {table_name}"
            df = pd.read_sql_query(query, conn)
            print(f"表 {table_name} 的数据：")
            print(f"  行数：{len(df)}")
            print(f"  列数：{len(df.columns)}")
            print(f"  列名：{list(df.columns)}")
            
            # 如果需要按Unicode排序
            if sort_by_unicode:
                print(f"按Unicode编码排序表 {table_name} 的字段名...")
                df = df.sort_index(axis=1)
                print(f"  排序后的列名：{list(df.columns)}")
            
            # 将数据写入Excel工作表
            print(f"将表 {table_name} 的数据写入Excel工作表...")
            df.to_excel(writer, sheet_name=table_name, index=False)
            print(f"成功将表 {table_name} 转换为Excel工作表")
        
        # 保存Excel文件
        print("保存Excel文件...")
        # 使用正确的保存方法
        writer.close()
        conn.close()
        
        print(f"成功将SQLite数据库转换为Excel文件：{excel_path}")
        return True
        
    except Exception as e:
        print(f"错误：转换SQLite到Excel时发生错误 - {str(e)}")
        import traceback
        traceback.print_exc()
        return False


def calculate_excel_functions(df):
    """
    计算Excel中的函数，返回计算后的DataFrame
    
    参数：
    df: 包含Excel函数的DataFrame
    
    返回：
    DataFrame: 计算后的DataFrame
    """
    try:
        # 创建一个临时Excel文件来计算函数
        temp_excel = "temp_calculate_functions.xlsx"
        
        # 将DataFrame写入临时Excel
        df.to_excel(temp_excel, index=False)
        
        # 重新加载Excel并计算函数
        wb = load_workbook(temp_excel)
        ws = wb.active
        
        # 遍历所有单元格，重新计算函数
        for row in ws.iter_rows(min_row=2, min_col=1):
            for cell in row:
                if cell.data_type == 'f':  # 如果是公式
                    cell.value = cell.value  # 重新设置值以触发计算
        
        # 保存计算后的Excel
        wb.save(temp_excel)
        wb.close()
        
        # 重新读取计算后的Excel
        calculated_df = pd.read_excel(temp_excel, engine='openpyxl')
        
        # 删除临时文件
        if os.path.exists(temp_excel):
            os.remove(temp_excel)
        
        return calculated_df
        
    except Exception as e:
        print(f"错误：计算Excel函数时发生错误 - {str(e)}")
        return df  # 返回原始DataFrame


def excel_to_sqlite(excel_path, db_path, table_name=None, sheet_name=0, calculate_functions=True):
    """
    将Excel文件转换为SQLite数据库表
    
    参数：
    excel_path: Excel文件路径
    db_path: 输出SQLite数据库文件路径
    table_name: 输出表名，None表示使用Excel工作表名
    sheet_name: Excel工作表名称或索引
    calculate_functions: 是否计算Excel中的函数
    
    返回：
    bool: 转换是否成功
    """
    try:
        print(f"开始Excel到SQLite的转换...")
        print(f"输入Excel文件：{excel_path}")
        print(f"输出SQLite文件：{db_path}")
        print(f"表名：{table_name}")
        print(f"工作表：{sheet_name}")
        print(f"是否计算函数：{calculate_functions}")
        
        # 检查Excel文件是否存在
        if not os.path.exists(excel_path):
            print(f"错误：Excel文件 {excel_path} 不存在！")
            return False
            
        # 检查Excel文件是否可读
        if not os.access(excel_path, os.R_OK):
            print(f"错误：无法读取Excel文件 {excel_path}！")
            return False
            
        # 检查输出目录是否存在
        output_dir = os.path.dirname(db_path)
        if output_dir and not os.path.exists(output_dir):
            try:
                os.makedirs(output_dir)
                print(f"已创建输出目录：{output_dir}")
            except Exception as e:
                print(f"错误：无法创建输出目录 {output_dir} - {str(e)}")
                return False
                
        # 读取Excel文件
        print(f"读取Excel文件中的数据...")
        df = pd.read_excel(excel_path, sheet_name=sheet_name, dtype=str)
        print(f"Excel文件数据：")
        print(f"  行数：{len(df)}")
        print(f"  列数：{len(df.columns)}")
        print(f"  列名：{list(df.columns)}")
        
        # 如果需要计算函数
        if calculate_functions:
            print("计算Excel中的函数...")
            df = calculate_excel_functions(df)
            print(f"函数计算后的数据：")
            print(f"  行数：{len(df)}")
            print(f"  列数：{len(df.columns)}")
        
        # 确定表名
        if table_name is None:
            print("确定输出表名...")
            # 获取工作表名
            xl = pd.ExcelFile(excel_path)
            sheet_names = xl.sheet_names
            print(f"Excel文件中的工作表：{sheet_names}")
            table_name = sheet_names[sheet_name] if isinstance(sheet_name, int) else sheet_name
            print(f"使用工作表名作为表名：{table_name}")
        
        # 清理表名（移除特殊字符）
        print(f"清理表名：{table_name}...")
        original_table_name = table_name
        table_name = ''.join(c if c.isalnum() or c == '_' else '_' for c in table_name)
        if not table_name[0].isalpha():
            table_name = 'table_' + table_name
        
        if table_name != original_table_name:
            print(f"表名已清理为：{table_name}")
        
        # 连接SQLite数据库
        print(f"连接SQLite数据库：{db_path}...")
        conn = sqlite3.connect(db_path)
        
        # 将DataFrame写入SQLite
        print(f"将数据写入SQLite表 {table_name}...")
        df.to_sql(table_name, conn, if_exists='replace', index=False)
        
        conn.close()
        
        print(f"成功将Excel工作表转换为SQLite表：{table_name}")
        print(f"成功将Excel文件转换为SQLite数据库：{db_path}")
        return True
        
    except Exception as e:
        print(f"错误：转换Excel到SQLite时发生错误 - {str(e)}")
        import traceback
        traceback.print_exc()
        return False


def get_db_tables(db_path):
    """
    获取SQLite数据库中的所有表名
    
    参数：
    db_path: SQLite数据库文件路径
    
    返回：
    list: 表名列表
    """
    try:
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%';")
        tables = [table[0] for table in cursor.fetchall()]
        conn.close()
        return tables
    except Exception as e:
        print(f"错误：获取表名时发生错误 - {str(e)}")
        return []


def get_table_columns(db_path, table_name):
    """
    获取SQLite表的所有字段名
    
    参数：
    db_path: SQLite数据库文件路径
    table_name: 表名
    
    返回：
    list: 字段名列表
    """
    try:
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        cursor.execute(f"PRAGMA table_info({table_name});")
        columns = [column[1] for column in cursor.fetchall()]
        conn.close()
        return columns
    except Exception as e:
        print(f"错误：获取字段名时发生错误 - {str(e)}")
        return []


if __name__ == "__main__":
    # 示例用法
    import argparse
    
    parser = argparse.ArgumentParser(description='SQLite与Excel相互转换工具')
    parser.add_argument('mode', choices=['sqlite2excel', 'excel2sqlite'], help='转换模式')
    parser.add_argument('input_file', help='输入文件路径')
    parser.add_argument('output_file', help='输出文件路径')
    parser.add_argument('--tables', nargs='*', help='要转换的表名列表（仅sqlite2excel模式）')
    parser.add_argument('--sheet', help='要转换的工作表名或索引（仅excel2sqlite模式）', default=0)
    parser.add_argument('--table_name', help='输出表名（仅excel2sqlite模式）')
    parser.add_argument('--calculate_functions', action='store_true', help='计算Excel函数（仅excel2sqlite模式）')
    parser.add_argument('--sort_unicode', action='store_true', help='按Unicode排序字段（仅sqlite2excel模式）')
    
    args = parser.parse_args()
    
    if args.mode == 'sqlite2excel':
        # SQLite转Excel
        success = sqlite_to_excel(args.input_file, args.output_file, args.tables, args.sort_unicode)
        if success:
            print("转换完成！")
        else:
            print("转换失败！")
            exit(1)
    else:
        # Excel转SQLite
        success = excel_to_sqlite(args.input_file, args.output_file, args.table_name, args.sheet, args.calculate_functions)
        if success:
            print("转换完成！")
        else:
            print("转换失败！")
            exit(1)